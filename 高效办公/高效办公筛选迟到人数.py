from openpyxl import load_workbook,Workbook

wb = load_workbook('10月考勤统计.xlsx')

ws = wb.active

late_header = []
for cell in ws[1]:
    late_header.append(cell.value)

new_wb = Workbook()
new_ws = new_wb.active

new_ws.append(late_header)

for row in ws.iter_rows(min_row = 2,values_only = True):
    name = row[1]
    department = row[2]
    time = row[3]
    number = row[4]
    if department == '人力资源部':
        if float(number) != 0 and time > 45 and number > 3:
            print('人力资源部的{}迟到了{}分钟，迟到了{}次'.format(name, time, number))
            new_ws.append(row)
new_wb.save('10月人力资源部迟到名单.xlsx')