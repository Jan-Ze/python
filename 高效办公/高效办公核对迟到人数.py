from openpyxl import load_workbook

# 打开工作簿【11月考勤统计.xlsx】，获取活动工作表
wb = load_workbook('11月考勤统计.xlsx')
ws = wb.active

# 创建迟到人员字典
late_p = {}

# 循环读取除表头外的表格数据
for row in ws.iter_rows(min_row = 2, values_only = True):
    ID = row[0]
    # 取出员工工号
    number = row[4]
    # 取出迟到次数
    late_p[ID] = number
    # 将信息添加入字典，字典格式为{'员工工号': '迟到次数'}


# 打开工作簿【迟到次数月度统计（11月更新）.xlsx】，获取活动工作表
monthly_wb = load_workbook('迟到次数月度统计（11月更新）.xlsx')
monthly_ws = monthly_wb.active

# 循环读取出表头外的表格数据
for monthly_row in monthly_ws.iter_rows(min_row = 3,values_only = True):
    ID =  monthly_row[0]
    # 取出员工工号
    number = monthly_row[13]
    # 取出十一月份的迟到次数
    if late_p[ID] != number:
    # 匹配迟到次数是否相等
    
        print('工号{}迟到情况不匹配，请核查后更新'.format(ID))