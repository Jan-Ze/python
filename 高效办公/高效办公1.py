"""题目要求
1. 通过代码，打开material文件夹下的文件practice2.xlsx，获取下半年公司名单工作表。
2. 打印第四列（D列）除表头部门外的所有数据。
3. 将原有的值全部修改为战略储备部。
4. 将结果保存为practice2_result.xlsx。
"""
from openpyxl import load_workbook

# 打开【practice2.xlsx】工作簿
staff_wb = load_workbook('practice2.xlsx')
# 按表名取表
staff_ws = staff_wb['下半年公司名单']

# 循环获取第四列（D列）的所有单元格对象
for col_cell in staff_ws['D']:
        if col_cell.value == '部门':
            continue
        col_cell.value = '傻逼'
# 将结果保存为【'practice2_result.xlsx'】
staff_wb.save('practice2_result.xlsx')

staff = load_workbook('practice2_result.xlsx')
staff_w = staff['上半年公司名单']
'''
for cell in staff_w.iter_rows(min_row = 1,max_row = 10,min_col=1,max_col=4,values_only = True):
    print(cell)
    '''